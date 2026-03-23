from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def parse_amount(value: object) -> float | None:
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return None

    cleaned = (
        text.replace("TND", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )

    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_duration_to_minutes(value: object) -> float | None:
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return None

    hours_match = re.search(r"(\d+)\s*h", text, flags=re.IGNORECASE)
    mins_match = re.search(r"(\d+)\s*min", text, flags=re.IGNORECASE)

    if not hours_match and not mins_match:
        return None

    hours = int(hours_match.group(1)) if hours_match else 0
    minutes = int(mins_match.group(1)) if mins_match else 0
    return float(hours * 60 + minutes)


def load_from_html(input_path: Path) -> pd.DataFrame:
    tables = pd.read_html(str(input_path))
    if not tables:
        raise ValueError(f"No table found in HTML file: {input_path}")

    expected = {
        "ID",
        "Username",
        "Firstname",
        "Lastname",
        "Duration",
        "Usage",
        "Ord.& Trans.",
        "USB Data",
        "Total Amount",
    }

    selected = None
    for table in tables:
        if isinstance(table.columns, pd.MultiIndex):
            table.columns = [" ".join([str(x) for x in tup if str(x) != "nan"]).strip() for tup in table.columns]
        cols = {str(c).strip() for c in table.columns}
        if expected.issubset(cols):
            selected = table.copy()
            break

    if selected is None:
        selected = tables[0].copy()

    selected = selected.fillna("")
    for col in expected:
        if col not in selected.columns:
            selected[col] = ""

    return selected[[
        "ID",
        "Username",
        "Firstname",
        "Lastname",
        "Duration",
        "Usage",
        "Ord.& Trans.",
        "USB Data",
        "Total Amount",
    ]]


def load_from_xls(input_path: Path) -> pd.DataFrame:
    df = pd.read_excel(str(input_path), header=None)

    column_mapping = {
        0: "ID",
        5: "Username",
        8: "Firstname",
        13: "Lastname",
        14: "Duration",
        22: "Usage",
        27: "Ord.& Trans.",
        28: "USB Data",
        30: "Total Amount",
    }

    df_fixed = df.iloc[5:, list(column_mapping.keys())].copy()
    df_fixed.columns = list(column_mapping.values())
    df_fixed = df_fixed.dropna(how="all")
    return df_fixed.fillna("")


def normalize(df_fixed: pd.DataFrame) -> pd.DataFrame:
    df_fixed = df_fixed.copy().fillna("")
    df_fixed = df_fixed.map(
        lambda x: str(x).strip() if isinstance(x, str) else x
    )
    df_fixed = df_fixed[df_fixed.notna().any(axis=1)]

    df_fixed = df_fixed[df_fixed["Username"].astype(str).str.lower() != "username"]

    for col in ["Usage", "Ord.& Trans.", "USB Data", "Total Amount"]:
        df_fixed[col] = df_fixed[col].apply(parse_amount)

    df_fixed["Duration (min)"] = df_fixed["Duration"].apply(parse_duration_to_minutes)

    ordered = [
        "ID",
        "Username",
        "Firstname",
        "Lastname",
        "Duration",
        "Duration (min)",
        "Usage",
        "Ord.& Trans.",
        "USB Data",
        "Total Amount",
    ]
    return df_fixed[ordered].reset_index(drop=True)


def format_excel(output_path: Path) -> None:
    wb = load_workbook(str(output_path))
    ws = wb.active

    for idx, col in enumerate(ws.columns, 1):
        max_length = 0
        letter = get_column_letter(idx)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[letter].width = min(max_length + 2, 50)

    wb.save(str(output_path))


def pick_default_input(project_root: Path) -> Path:
    candidates = [
        project_root / "data" / "memeber DATA 01-09-2025.xls",
        project_root / "members" / "member_week.xls",
        project_root / "data" / "test.html",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("No members input file found in expected locations.")


def main() -> None:
    project_root = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(
        description="Members report cleaner (prefers HTML, falls back to XLS)."
    )
    parser.add_argument("--input", type=str, default="", help="Input report path (.html/.xls/.xlsx)")
    parser.add_argument("--output", type=str, default="", help="Output .xlsx path")
    args = parser.parse_args()

    input_path = Path(args.input) if args.input else pick_default_input(project_root)
    ext = input_path.suffix.lower()

    if ext in {".html", ".htm"}:
        source_df = load_from_html(input_path)
    elif ext in {".xls", ".xlsx"}:
        source_df = load_from_xls(input_path)
    else:
        raise ValueError("Unsupported input extension. Use .html, .htm, .xls or .xlsx")

    cleaned = normalize(source_df)

    if args.output:
        output_path = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = project_root / "members" / f"members_fixed_{timestamp}.xlsx"

    cleaned.to_excel(str(output_path), index=False)
    format_excel(output_path)

    print(f"Input file: {input_path}")
    print(f"Rows cleaned: {len(cleaned)}")
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
