from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def parse_amount(value: object) -> float | None:
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return None

    cleaned = (
        text.replace("TND", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )

    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_number(value: object) -> float | None:
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return None

    cleaned = text.replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def load_from_html(input_path: Path) -> pd.DataFrame:
    tables = pd.read_html(str(input_path))
    if not tables:
        raise ValueError(f"No table found in HTML file: {input_path}")

    expected = {
        "Cashier",
        "Date",
        "Item Name",
        "Category",
        "In/Out",
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Comment",
    }

    selected = None
    for table in tables:
        if isinstance(table.columns, pd.MultiIndex):
            table.columns = [" ".join([str(x) for x in tup if str(x) != "nan"]).strip() for tup in table.columns]
        cols = {str(c).strip() for c in table.columns}
        if expected.issubset(cols):
            selected = table.copy()
            break

    if selected is None:
        selected = tables[0].copy()

    selected = selected.fillna("")
    for col in expected:
        if col not in selected.columns:
            selected[col] = ""

    return selected[[
        "Cashier",
        "Date",
        "Item Name",
        "Category",
        "In/Out",
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Comment",
    ]]


def load_from_xls(input_path: Path) -> pd.DataFrame:
    df = pd.read_excel(str(input_path), header=None)
    df = df.dropna(how="all")

    rows = []
    for _, row in df.iterrows():
        vals = [v for v in row if pd.notna(v)]
        rows.append(vals)

    df2 = pd.DataFrame(rows)
    df2 = df2[df2[0] != "Stock Reports"]
    df2 = df2[df2[0] != "Cashier"]
    df2 = df2[~df2[0].astype(str).str.match(r"\d{2}/\d{2}/\d{4}")]

    df_fixed = pd.DataFrame(
        {
            "Cashier": df2[0],
            "Date": df2[1],
            "Item Name": df2[2],
            "Category": df2[3],
            "In/Out": df2[4],
            "Quantity": df2[5],
            "Unit Price": df2[6],
            "Total Amount": df2[7],
            "Comment": df2[8],
        }
    ).reset_index(drop=True)

    return df_fixed.fillna("")


def normalize(df_fixed: pd.DataFrame) -> pd.DataFrame:
    df_fixed = df_fixed.copy().fillna("")

    df_fixed = df_fixed[df_fixed["Cashier"] != "Cashier"]
    df_fixed = df_fixed[
        ~df_fixed["Cashier"].astype(str).str.contains(r"\d{2}/\d{2}/\d{4}", na=False)
    ]

    df_fixed["Quantity"] = df_fixed["Quantity"].apply(parse_number)
    df_fixed["Unit Price"] = df_fixed["Unit Price"].apply(parse_amount)
    df_fixed["Total Amount"] = df_fixed["Total Amount"].apply(parse_amount)

    ordered = [
        "Cashier",
        "Date",
        "Item Name",
        "Category",
        "In/Out",
        "Quantity",
        "Unit Price",
        "Total Amount",
        "Comment",
    ]
    return df_fixed[ordered].reset_index(drop=True)


def format_excel(output_path: Path) -> None:
    wb = load_workbook(str(output_path))
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(str(output_path))


def pick_default_input(project_root: Path) -> Path:
    candidates = [
        project_root / "data" / "Stock DATA 01-09-2025.xls",
        project_root / "stock" / "stock_lastmonth_xls.xls",
        project_root / "data" / "test.html",
    ]
    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("No stock input file found in expected locations.")


def main() -> None:
    project_root = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(
        description="Stock report cleaner (prefers HTML, falls back to XLS)."
    )
    parser.add_argument("--input", type=str, default="", help="Input report path (.html/.xls/.xlsx)")
    parser.add_argument("--output", type=str, default="", help="Output .xlsx path")
    args = parser.parse_args()

    input_path = Path(args.input) if args.input else pick_default_input(project_root)
    ext = input_path.suffix.lower()

    if ext in {".html", ".htm"}:
        source_df = load_from_html(input_path)
    elif ext in {".xls", ".xlsx"}:
        source_df = load_from_xls(input_path)
    else:
        raise ValueError("Unsupported input extension. Use .html, .htm, .xls or .xlsx")

    cleaned = normalize(source_df)

    if args.output:
        output_path = Path(args.output)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = project_root / "stock" / f"stock_fixed_{timestamp}.xlsx"

    cleaned.to_excel(str(output_path), index=False)
    format_excel(output_path)

    print(f"Input file: {input_path}")
    print(f"Rows cleaned: {len(cleaned)}")
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()